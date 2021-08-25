import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import time
import datetime
import matplotlib
from scipy import interpolate

#
# User settings
#

#   directory of TOUGH results
fdir = '/Users/zli/Documents/HFTS/results_fracture_network/'
#   directory of field data
prodir = '/Users/zli/Documents/HFTS/data/4SM and 4SU Production Tracking 01-22-20.xlsx'

datdir = ['upslo2_gor2800_1500/','upslo2_gor2800_hf1500/','upslo2_gor2800_sf1500/','upslo2_gor2800_hfsf1500/']
flabel = ['No closure','HF closed','SF closed','HF+SF closed']
fname = ['Conx_Time_Series','Conx_Time_Series','Conx_Time_Series',
            'Conx_Time_Series','Conx_Time_Series','Conx_Time_Series',
            'Conx_Time_Series','Conx_Time_Series','Conx_Time_Series']

ylab = ['Water','Gas','Oil']

plot_cumu = True
plot_data = True
savefig = False
savename = 'fig_fracture_closure'

dlabel = ['1714M']
tfig = 4
fs = 7
ls = ['-','-','--','-','--','-',':',':',':']
co = ['r','g','g','b','c','k','b','b','g']
ymax = [1.0,0.5,0.5]
ind = [' ','(a)','(b)','(c)','(d)','(e)','(f)']

dt_sim = 2400.0
n_stage = 185
# coeff = n_stage / 5.0 * np.ones(len(datdir))
coeff = n_stage * np.ones(len(datdir))
# coeff[-1] = coeff[-1] / 5.0


# unit conversion
kg2bbl = 0.00839
kg2mcf = 0.023

kg2bbl_water = 6.29 / 1e3
kg2bbl_oil = 6.29 / 8e2
unit_conv = [kg2bbl_water, kg2mcf, kg2bbl_oil]

bc = {
        'isheet'    :   0,
        'period'    :   ['20160110', '20200104'],
        'col'       :   [0,3,2,1,7,8],
        'units'     :   ['none', 'bbl2kgw', 'mcf2kg', 'bbl2kgo', 'psi2Pa', 'F2C']
}

def loadFile(datdir, fname):
    """ Load TOUGH output data file

        Returns:
            data (nparray)  : [time, water rate, gas rate]
    """
    time = []
    water = []
    gas = []
    oil = []
    # Open file
    fid = open(fdir+datdir+fname,"r")
    title = str.split(fid.readline())
    # Read the rest of the datafile
    tt = -1
    tbase = 0
    while True:
        line = str.split(fid.readline())
        if not line or len(line) < 2:
            break
        else:
            try:
                float(line[0])
            except ValueError:
                line = str.split(fid.readline())
                tbase = time[-1]


            time.append(float(line[0]) + tbase)
            water.append(float(line[2]))
            gas.append(float(line[1]))
            oil.append(float(line[3]))
    fid.close()
    # convert list to numpy array
    N = len(time)
    data = np.zeros((N,4), dtype=float)
    data[:,0] = np.array(time)
    data[:,1] = np.array(water)
    data[:,2] = np.array(gas)
    data[:,3] = np.array(oil)
    return data

def excel_date(date1, base=36525):
    temp = time.mktime(datetime.datetime.strptime('19991230','%Y%m%d').timetuple())
    return base + (date1 - temp)/86400

def unitConv(value, func):
    """ Unit conversion functions """
    if func == 'none':
        cfun = lambda x: x
    elif func == 'psi2Pa':
        cfun = lambda x: x * 6894.76
    elif func == 'd2sec':
        cfun = lambda x: x * 86400.0
    elif func == 'F2C':
        cfun = lambda x: (x - 32.0) * 5.0 / 9.0
    elif func == 'bbl2kgo':
        cfun = lambda x: x * 0.158987 * 790
    elif func == 'bbl2kgw':
        cfun = lambda x: x * 0.158987 * 990
    elif func == 'mcf2kg':
        cfun = lambda x: x * 28.3168 * 0.71
    else:
        raise ValueError('Undefined unit conversion!')
    return cfun(value)

def readBC(fname, bc, sheet_label):
    """ Read time-variable boundary condition from excel

    This function is designed specifically for HFTS Production Data (in excel).
    For other types of input data, should define other functions.

    Args:
        fname (str)     : Excel file name to read
        bc (dict)       : Dict containing bc settings, which includes:
            isheet (int)    : Index of sheet to read
            period (list)   : [start end] date in the format of yyyymmdd
            col (list)      : Columns in excel to be read
            units (list)    : Unit conversion function for each column

    Returns:
        data (nparray)  : Data extracted from excel file within the range of 'period'
    """
    # top 10 lines are not date numbers, so it is removed
    offset = 10
    # load excel file
    # datafile = xlrd.open_workbook(fname)
    datafile = load_workbook(fname)
    sheet = datafile[sheet_label]
    allCells = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])

    # allCells is zero-indexed
    ar = allCells[offset:1488,:10]
    # sheet = datafile.sheet_by_index(bc['isheet'])
    # times = sheet.col_values(0)
    # del times[1488:]
    # del times[0:offset]
    # del times[780:]
    # times = np.array(times, dtype=float)
    times = []
    for ii in range(np.shape(ar)[0]):
        times.append(excel_date(time.mktime(ar[ii,0].timetuple())))
    times = np.array(times)
    # find row index corresponds to the desired time period
    t0 = time.mktime(datetime.datetime.strptime(bc['period'][0],'%Y%m%d').timetuple())
    t1 = time.mktime(datetime.datetime.strptime(bc['period'][1],'%Y%m%d').timetuple())

    ind0 = np.where(abs(times - excel_date(t0)) < 0.1)[0][0] + offset - 1
    ind1 = np.where(abs(times - excel_date(t1)) < 0.1)[0][0] + offset - 1
    # extract data into numpy array
    data = np.zeros((ind1-ind0,len(bc['col'])))
    counter = 0
    for kk in range(ind0, ind1):
        for ii in range(len(bc['col'])):
            # if type(sheet.cell_value(kk,bc['col'][ii])) is float:
                # data[counter,ii] = sheet.cell_value(kk,bc['col'][ii])
            if type(ar[kk,bc['col'][ii]]) is float:
                data[counter,ii] = ar[kk,bc['col'][ii]]
            elif type(ar[kk,bc['col'][ii]]) is int:
                data[counter,ii] = float(ar[kk,bc['col'][ii]])
            else:
                data[counter,ii] = np.nan
            data[counter,ii] = unitConv(data[counter,ii], bc['units'][ii])
        counter += 1
    data[:,0] = times[ind0:ind1]
    data[:,0] -= data[0,0]
    return data

""" Execution """
simu = {}
for ii in range(len(datdir)):
    simu[str(ii)] = loadFile(datdir[ii], fname[ii])
    dt = (simu[str(ii)][1:,0] - simu[str(ii)][:-1,0]) * 86400.0
    simu[str(ii)][1:,1] = simu[str(ii)][1:,1] * coeff[ii]
    simu[str(ii)][1:,2] = simu[str(ii)][1:,2] * coeff[ii]
    simu[str(ii)][1:,3] = simu[str(ii)][1:,3] * coeff[ii]

if plot_data:
    data = readBC(prodir, bc, '171 4SM')
    data[:,1:] = data[:,1:] / 86400.0
    print('Mean BHP = ',np.nanmean(data[:,4]) * 86400 / 1e6, ' MPa')

    datau = readBC(prodir, bc, '171 4SU')
    datau[:,1:] = datau[:,1:] / 86400.0

# get cumulative flow rates
if plot_data is True:
    data_cumu = np.zeros((np.shape(data)[0],4))
    datau_cumu = np.zeros((np.shape(datau)[0],4))
    for ii in range(1,np.shape(data)[0]):
        data_cumu[ii,0] = data[ii,0]
        datau_cumu[ii,0] = datau[ii,0]
        for jj in range(1,4):
            # data_cumu[ii,jj] = np.nansum(data[1:ii,jj] * (data[1:ii,0]-data[0:ii-1,0])) * 86400
            data_cumu[ii,jj] = np.nansum(data[1:ii,jj]) * 86400
            datau_cumu[ii,jj] = np.nansum(datau[1:ii,jj]) * 86400
simu_cumu = {}
for ff in range(len(datdir)):
    simu_cumu[str(ff)] = np.zeros((np.shape(simu[str(ff)])[0],4))
    for ii in range(1,np.shape(simu_cumu[str(ff)])[0]):
        simu_cumu[str(ff)][ii,0] = simu[str(ff)][ii,0]
        for jj in range(1,4):
            simu_cumu[str(ff)][ii,jj] = np.nansum(simu[str(ff)][1:ii,jj] * (simu[str(ff)][1:ii,0]-simu[str(ff)][0:ii-1,0])) * 86400.0

# interpolate
intp = []
intp_cumu = []
tvec = np.linspace(1, 100.0, 1000)
for ii in range(len(datdir)):
    intp.append([])
    intp_cumu.append([])
    for jj in range(3):
        fintp1 = interpolate.interp1d(simu[str(ii)][:,0], simu[str(ii)][:,jj+1])
        fintp2 = interpolate.interp1d(simu_cumu[str(ii)][:,0], simu_cumu[str(ii)][:,jj+1])
        intp[ii].append(fintp1(tvec))
        intp_cumu[ii].append(fintp2(tvec))

# Calculate average
simu_avg = np.nanmean(intp, axis=0)
simu_avg_cumu = np.nanmean(intp_cumu, axis=0)



ifig = 1
font = {'family' : 'Arial',
        'size'   : fs}
matplotlib.rc('font', **font)
cm = 1.0 / 2.54


ifig = 1
fig = plt.figure(1, figsize=[18*cm,7*cm])
for ff in range(3):
    plt.subplot(1,3,ifig)
    ax = fig.gca()
    pos1 = ax.get_position()
    pos2 = [pos1.x0 + 0.045*(ff-1), pos1.y0+0.015, pos1.width*1.1, pos1.height]
    ax.set_position(pos2)

    print('    --------    ')
    if plot_data is True:
        plt.plot(data[:,0]+0.5, data[:,ff+1], label='Field data', color='k', linewidth=1.0, linestyle=':')
    for ii in range(len(datdir)):
        plt.plot(simu[str(ii)][1:,0], abs(simu[str(ii)][1:,ff+1]), label=flabel[ii],
            color=co[ii], linestyle=ls[ii], linewidth=0.7)
        # if ii == len(datdir)-1:
        #     plt.plot(tvec, simu_avg[ff,:], label='avg', color='g', linestyle=ls[ii], linewidth=0.7)

    plt.ylabel(ylab[ff]+' rate [kg/s]')
    plt.xlabel('Days')
    plt.ticklabel_format(style='sci', axis='y',scilimits=(0,0))
    plt.ticklabel_format(useMathText=True)
    plt.xlim([0,1500])
    if ff == 2:
        plt.ylim([0,1])
        # plt.ylim([0,1e7])
        plt.legend(loc='upper right')
        # plt.annotate(ind[ifig], (50,0.9*2e7), color='k', fontsize=fs)
    elif ff == 1:
        plt.ylim([0,0.4])
        # plt.annotate(ind[ifig], (50,0.9*0.5e7), color='k', fontsize=fs)
    else:
        plt.ylim([0,2])
        # plt.annotate(ind[ifig], (50,0.9*3.5e7), color='k', fontsize=fs)
    ifig += 1


if savefig:
    plt.savefig(savename+'.eps', format='eps')

ifig = 1
fig = plt.figure(2, figsize=[18*cm,7*cm])
for ff in range(3):
    plt.subplot(1,3,ifig)
    ax = fig.gca()
    pos1 = ax.get_position()
    pos2 = [pos1.x0 + 0.045*(ff-1), pos1.y0+0.015, pos1.width*1.1, pos1.height]
    ax.set_position(pos2)

    if plot_data is True:
        plt.plot(data_cumu[:,0]+0.5, data_cumu[:,ff+1], label='Field data', color='k', linewidth=1.0, linestyle=':')
    for ii in range(len(datdir)):
        plt.plot(simu_cumu[str(ii)][1:,0], abs(simu_cumu[str(ii)][1:,ff+1]), label=flabel[ii],
            color=co[ii], linestyle=ls[ii], linewidth=0.7)
        # if ii == len(datdir)-1:
        #     plt.plot(tvec, simu_avg_cumu[ff,:], label='avg', color='g', linestyle=ls[ii], linewidth=0.7)

    plt.ylabel('Total '+ylab[ff]+'[kg]')
    plt.xlabel('Days')
    plt.ticklabel_format(style='sci', axis='y',scilimits=(0,0))
    plt.ticklabel_format(useMathText=True)
    plt.xlim([0,1500])
    if ff == 1:
        # plt.ylim([0,1])
        # plt.ylim([0,1e7])
        plt.legend(loc='upper left')
        # plt.annotate(ind[ifig], (50,0.9*2e7), color='k', fontsize=fs)
    # elif ff == 1:
        # plt.ylim([0,0.4])
        # plt.annotate(ind[ifig], (50,0.9*0.5e7), color='k', fontsize=fs)
    # else:
        # plt.ylim([0,2])
        # plt.annotate(ind[ifig], (50,0.9*3.5e7), color='k', fontsize=fs)
    ifig += 1

if savefig:
    plt.savefig(savename+'_cumu.eps', format='eps')

plt.show()
